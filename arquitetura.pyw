import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from ttkthemes import ThemedStyle
from datetime import datetime
import openpyxl

def obter_projetos_existentes():
    planilha = openpyxl.load_workbook("registro.xlsx")
    planilha_ativa = planilha.active

    projetos = []
    for linha in range(1, planilha_ativa.max_row + 1):
        projeto = planilha_ativa.cell(row=linha, column=1).value
        if projeto:
            projetos.append(projeto)

    return projetos


def verificar_projeto_existente(projeto):
    projetos_existentes = obter_projetos_existentes()
    return projeto in projetos_existentes


def validar_data(data):
    try:
        datetime.strptime(data, "%d/%m/%Y")
        return True
    except ValueError:
        return False


def registrar_entrada1():
    projeto = projeto_selecionado.get()

    if projeto == "":
        messagebox.showerror("Erro", "Por favor, selecione um projeto.")
        return

    if not verificar_projeto_existente(projeto):
        messagebox.showerror("Erro", "O projeto selecionado não existe.")
        return

    planilha = openpyxl.load_workbook("registro.xlsx")
    planilha_ativa = planilha.active

    for linha in range(1, planilha_ativa.max_row + 1):
        if planilha_ativa.cell(row=linha, column=1).value == projeto:
            if planilha_ativa.cell(row=linha, column=3).value is not None:
                messagebox.showerror("Erro", "A 1ª entrada já foi registrada para o projeto.")
                return

            hora_entrada = datetime.now().strftime("%H:%M:%S")
            planilha_ativa.cell(row=linha, column=3).value = hora_entrada
            planilha.save("registro.xlsx")
            messagebox.showinfo("Sucesso", "1ª Entrada registrada com sucesso!")
            return

    messagebox.showerror("Erro", "Projeto não encontrado.")


def registrar_entrada2():
    projeto = projeto_selecionado.get()

    if projeto == "":
        messagebox.showerror("Erro", "Por favor, selecione um projeto.")
        return

    if not verificar_projeto_existente(projeto):
        messagebox.showerror("Erro", "O projeto selecionado não existe.")
        return

    planilha = openpyxl.load_workbook("registro.xlsx")
    planilha_ativa = planilha.active

    for linha in range(1, planilha_ativa.max_row + 1):
        if planilha_ativa.cell(row=linha, column=1).value == projeto:
            if planilha_ativa.cell(row=linha, column=5).value is not None:
                messagebox.showerror("Erro", "A 2ª entrada já foi registrada para o projeto.")
                return

            hora_entrada = datetime.now().strftime("%H:%M:%S")
            planilha_ativa.cell(row=linha, column=5).value = hora_entrada
            planilha.save("registro.xlsx")
            messagebox.showinfo("Sucesso", "2ª Entrada registrada com sucesso!")
            return

    messagebox.showerror("Erro", "Projeto não encontrado.")


def registrar_saida1():
    projeto = projeto_selecionado.get()

    if projeto == "":
        messagebox.showerror("Erro", "Por favor, selecione um projeto.")
        return

    if not verificar_projeto_existente(projeto):
        messagebox.showerror("Erro", "O projeto selecionado não existe.")
        return

    planilha = openpyxl.load_workbook("registro.xlsx")
    planilha_ativa = planilha.active

    for linha in range(1, planilha_ativa.max_row + 1):
        if planilha_ativa.cell(row=linha, column=1).value == projeto:
            if planilha_ativa.cell(row=linha, column=4).value is not None:
                messagebox.showerror("Erro", "A 1ª saída já foi registrada para o projeto.")
                return

            hora_saida = datetime.now().strftime("%H:%M:%S")
            planilha_ativa.cell(row=linha, column=4).value = hora_saida
            planilha.save("registro.xlsx")
            messagebox.showinfo("Sucesso", "1ª Saída registrada com sucesso!")
            return

    messagebox.showerror("Erro", "Projeto não encontrado.")


def registrar_saida2():
    projeto = projeto_selecionado.get()

    if projeto == "":
        messagebox.showerror("Erro", "Por favor, selecione um projeto.")
        return

    if not verificar_projeto_existente(projeto):
        messagebox.showerror("Erro", "O projeto selecionado não existe.")
        return

    planilha = openpyxl.load_workbook("registro.xlsx")
    planilha_ativa = planilha.active

    for linha in range(1, planilha_ativa.max_row + 1):
        if planilha_ativa.cell(row=linha, column=1).value == projeto:
            if planilha_ativa.cell(row=linha, column=6).value is not None:
                messagebox.showerror("Erro", "A 2ª saída já foi registrada para o projeto.")
                return

            hora_saida = datetime.now().strftime("%H:%M:%S")
            planilha_ativa.cell(row=linha, column=6).value = hora_saida
            planilha.save("registro.xlsx")
            messagebox.showinfo("Sucesso", "2ª Saída registrada com sucesso!")
            return

    messagebox.showerror("Erro", "Projeto não encontrado.")


def criar_projeto():
    nome_novo_projeto = nome_projeto.get()
    data_novo_projeto = data_projeto.get()

    if nome_novo_projeto == "":
        messagebox.showerror("Erro", "Por favor, digite um nome para o novo projeto.")
        return

    if data_novo_projeto == "":
        messagebox.showerror("Erro", "Por favor, selecione uma data para o novo projeto.")
        return

    if verificar_projeto_existente(nome_novo_projeto):
        messagebox.showerror("Erro", "Já existe um projeto com esse nome.")
        return

    if not validar_data(data_novo_projeto):
        messagebox.showerror("Erro", "A data está no formato incorreto. Utilize dd/mm/aaaa.")
        return

    planilha = openpyxl.load_workbook("registro.xlsx")
    planilha_ativa = planilha.active

    nova_linha = planilha_ativa.max_row + 1
    planilha_ativa.cell(row=nova_linha, column=1).value = nome_novo_projeto
    planilha_ativa.cell(row=nova_linha, column=2).value = data_novo_projeto

    planilha.save("registro.xlsx")
    projeto_selecionado.set(nome_novo_projeto)

    projetos_existentes = obter_projetos_existentes()
    combo_projeto['values'] = projetos_existentes

    nome_projeto.delete(0, 'end')
    data_projeto.delete(0, 'end')

    messagebox.showinfo("Sucesso", "Novo projeto criado com sucesso!")


def validar_data(data):
    try:
        datetime.strptime(data, "%d/%m/%Y")
        return True
    except ValueError:
        return False

janela = tk.Tk()
janela.title("Registro de Projetos")
janela.geometry("400x600")
janela.resizable(False, False)
janela.configure(bg="#708090")

style = ThemedStyle(janela)
style.set_theme("clam")
style.configure("TButton",
                font=("Arial", 12, "bold"),
                background="#B0C4DE",
                foreground="#333333",
                padding=10)
style.map("TButton",
          foreground=[('active', 'white')],
          background=[('active', '#007bff')])


# Ícone do programa
janela.iconbitmap("arquitetura.ico")

projetos_existentes = obter_projetos_existentes()

projeto_selecionado = tk.StringVar()
data_selecionada = tk.StringVar()

style = ttk.Style()
style.configure("TButton", font=("Arial", 12))

titulo = ttk.Label(janela, text="Registro de Projetos", style="TLabel")
titulo.pack(pady=20)
style.configure("TLabel",
                font=("Arial", 20, "bold"),
                foreground="#F8F8FF", 
                background="#708090")  

label_nome = tk.Label(janela, text="Nome:")
label_nome.pack(pady=10)

nome_projeto = ttk.Entry(janela)
nome_projeto.pack()

label_data = tk.Label(janela, text="Data (dd/mm/aaaa):")
label_data.pack(pady=10)

data_projeto = ttk.Entry(janela, textvariable=data_selecionada)
data_projeto.pack()

btn_criar = ttk.Button(janela, text="Criar Projeto", command=criar_projeto)
btn_criar.pack(pady=10)

label_projeto = tk.Label(janela, text="Projeto:")
label_projeto.pack(pady=5)

combo_projeto = ttk.Combobox(janela, textvariable=projeto_selecionado, values=projetos_existentes)
combo_projeto.pack(pady=5)

style.configure("TButton", font=("Arial", 12), width=20,)

btn_entrada1 = ttk.Button(janela, text="Registrar 1ª Entrada", command=registrar_entrada1, style="TButton")
btn_entrada1.pack(pady=10)

btn_saida1 = ttk.Button(janela, text="Registrar 1ª Saída", command=registrar_saida1, style="TButton")
btn_saida1.pack(pady=10)

btn_entrada2 = ttk.Button(janela, text="Registrar 2ª Entrada", command=registrar_entrada2, style="TButton")
btn_entrada2.pack(pady=10)

btn_saida2 = ttk.Button(janela, text="Registrar 2ª Saída", command=registrar_saida2, style="TButton")
btn_saida2.pack(pady=10)

def destacar_botao(event):
    event.widget.configure(background="#FFD700", foreground="#000000", relief=tk.SOLID) 
    event.widget.config(cursor="hand2")  
def restaurar_botao(event):
    event.widget.configure(background="#1f1f1f", foreground="#FFFFFF", relief=tk.FLAT)  
    event.widget.config(cursor="") 

style = ttk.Style()
style.configure("TButton",
                font=("Arial", 12, "bold"),
                background="#191970", 
                foreground="#FFFFFF", 
                relief=tk.FLAT, 
                padding=10)

btn_criar = ttk.Button(janela, text="Criar Projeto", style="TButton")
btn_entrada1 = ttk.Button(janela, text="", style="TButton")
btn_saida1 = ttk.Button(janela, text="Registrar 1ª Saída", style="TButton")
btn_entrada2 = ttk.Button(janela, text="Registrar 2ª Entrada", style="TButton")
btn_saida2 = ttk.Button(janela, text="Registrar 2ª Saída", style="TButton")

btn_criar.bind("<Enter>", destacar_botao)
btn_criar.bind("<Leave>", restaurar_botao)
btn_entrada1.bind("<Enter>", destacar_botao)
btn_entrada1.bind("<Leave>", restaurar_botao)
btn_saida1.bind("<Enter>", destacar_botao)
btn_saida1.bind("<Leave>", restaurar_botao)
btn_entrada2.bind("<Enter>", destacar_botao)
btn_entrada2.bind("<Leave>", restaurar_botao)
btn_saida2.bind("<Enter>", destacar_botao)
btn_saida2.bind("<Leave>", restaurar_botao)

janela.mainloop()
